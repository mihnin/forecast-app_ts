import pandas as pd
import io
import logging
from typing import Dict, Any, Optional
from openpyxl.styles import PatternFill

logger = logging.getLogger(__name__)

def generate_excel_buffer(preds: pd.DataFrame, leaderboard: Optional[pd.DataFrame] = None,
                        static_train: Optional[pd.DataFrame] = None,
                        ensemble_info_df: Optional[pd.DataFrame] = None) -> io.BytesIO:
    """
    Формирует Excel-файл в памяти с листами:
      - Predictions
      - Leaderboard с подсветкой лучшей модели
      - StaticTrainFeatures
      - WeightedEnsembleInfo
    
    Args:
        preds: DataFrame с прогнозами
        leaderboard: Таблица лидеров
        static_train: DataFrame со статическими признаками
        ensemble_info_df: DataFrame с информацией об ансамбле
        
    Returns:
        BytesIO объект с Excel-файлом
    """
    excel_buffer = io.BytesIO()
    
    try:
        with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
            # Лист с предсказаниями
            if preds is not None:
                if isinstance(preds, pd.DataFrame):
                    preds.reset_index().to_excel(writer, sheet_name="Predictions", index=False)
                else:
                    # Если preds - список словарей (из API)
                    pd.DataFrame(preds).to_excel(writer, sheet_name="Predictions", index=False)
            
            # Лидерборд с подсветкой лучшей модели
            if leaderboard is not None:
                if isinstance(leaderboard, pd.DataFrame):
                    leaderboard_df = leaderboard
                else:
                    # Если leaderboard - список словарей (из API)
                    leaderboard_df = pd.DataFrame(leaderboard)
                
                leaderboard_df.to_excel(writer, sheet_name="Leaderboard", index=False)
                
                try:
                    sheet_lb = writer.sheets["Leaderboard"]
                    # Подсветка лучшей модели
                    fill_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    # Первая строка после заголовка
                    for col_idx in range(1, leaderboard_df.shape[1] + 1):
                        cell = sheet_lb.cell(row=2, column=col_idx)
                        cell.fill = fill_green
                except Exception as e:
                    logger.error(f"Ошибка при подсветке лучшей модели в Leaderboard: {e}")
            
            # Лист со статическими признаками
            if static_train is not None and not static_train.empty:
                static_train.to_excel(writer, sheet_name="StaticTrainFeatures", index=False)
            
            # Лист с информацией об ансамбле
            if ensemble_info_df is not None and (isinstance(ensemble_info_df, pd.DataFrame) and not ensemble_info_df.empty):
                ensemble_info_df.to_excel(writer, sheet_name="WeightedEnsembleInfo", index=False)
            
            # Если ensemble_info_df - словарь (из API)
            if ensemble_info_df is not None and isinstance(ensemble_info_df, dict):
                try:
                    models = ensemble_info_df.get("models", [])
                    weights = ensemble_info_df.get("weights", [])
                    
                    if models and weights and len(models) == len(weights):
                        ensemble_df = pd.DataFrame({"Model": models, "Weight": weights})
                        ensemble_df.to_excel(writer, sheet_name="WeightedEnsembleInfo", index=False)
                except Exception as e:
                    logger.error(f"Ошибка при сохранении информации об ансамбле: {e}")
    
    except Exception as e:
        logger.error(f"Ошибка при создании Excel-файла: {e}")
        raise
    
    return excel_buffer